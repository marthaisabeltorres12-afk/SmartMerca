from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt
from models.product import Product
from models.supplier import Supplier
from extensions import db
from datetime import datetime
import io, re

import_export_bp = Blueprint('import_export', __name__)

# ── EXPORTAR a Excel ──────────────────────────────────────────────────────
@import_export_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def exportar_excel():
    claims = get_jwt()
    if claims.get('role') not in ('admin', 'admin_tecnico'):
        return jsonify({'message': 'Sin permiso'}), 403
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'message': 'Instala openpyxl: pip install openpyxl'}), 500

    productos = Product.query.filter_by(is_active=True).order_by(Product.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos SmartMerca'

    # Estilos
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Encabezados
    headers = [
        'Nombre', 'Descripción', 'Precio', 'Stock', 'Stock Mínimo',
        'Categoría', 'Código de Barras', 'Proveedor', 'IVA (%)',
        'Gramaje Cantidad', 'Gramaje Unidad', 'Fecha Vencimiento',
        'Descuento (%)', 'Activo'
    ]
    col_widths = [30, 35, 12, 8, 12, 20, 18, 20, 8, 14, 14, 18, 12, 8]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Datos
    for row, p in enumerate(productos, 2):
        fill_alt = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid') \
            if row % 2 == 0 else PatternFill(fill_type=None)
        valores = [
            p.name,
            p.description or '',
            float(p.price),
            p.stock,
            p.min_stock or 5,
            p.category or '',
            p.barcode or '',
            (p.supplier.company_name or p.supplier.name) if p.supplier else '',
            p.iva_type or 19,
            float(p.gramaje_cantidad) if p.gramaje_cantidad else '',
            p.gramaje_unidad or '',
            str(p.expiry_date) if p.expiry_date else '',
            float(p.discount) if p.discount else 0,
            'Si' if p.is_active else 'No',
        ]
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border    = border
            cell.fill      = fill_alt
            cell.alignment = Alignment(vertical='center')
            if col == 3:  # precio
                cell.number_format = '#,##0'

    # Hoja de instrucciones
    ws2 = wb.create_sheet('Instrucciones')
    instrucciones = [
        ['INSTRUCCIONES PARA IMPORTAR PRODUCTOS'],
        [''],
        ['1. Llena la hoja "Productos SmartMerca" con tus productos'],
        ['2. Columnas OBLIGATORIAS: Nombre, Precio, Stock'],
        ['3. IVA: 0, 5 o 19 (porcentaje sin %)'],
        ['4. Gramaje Unidad: kg, g, lb, ml, l, unidad'],
        ['5. Si el producto ya existe (mismo nombre), se ACTUALIZA'],
        ['6. Si es nuevo, se CREA'],
        ['7. Código de barras: déjalo vacío si no tiene'],
        ['8. Descuento: porcentaje entre 0 y 100'],
        [''],
        ['EJEMPLO:'],
        ['Nombre',  'Precio', 'Stock', 'Categoría', 'IVA (%)', 'Código de Barras'],
        ['Arroz Diana 500g', '3500', '50', 'Granos', '0', '7702001234567'],
        ['Aceite Palma 1L', '8900', '30', 'Aceites', '19', ''],
        ['Banano', '2500', '40', 'Frutas y Verduras', '0', ''],
    ]
    for i, fila in enumerate(instrucciones, 1):
        for j, val in enumerate(fila, 1):
            cell = ws2.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = Font(bold=True, size=13, color='1E3A5F')
            if i == 13:
                cell.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 35

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'SmartMerca_Productos_{fecha}.xlsx'
    )


# ── IMPORTAR desde Excel ──────────────────────────────────────────────────
@import_export_bp.route('/import/excel', methods=['POST'])
@jwt_required()
def importar_excel():
    claims = get_jwt()
    if claims.get('role') not in ('admin', 'admin_tecnico'):
        return jsonify({'message': 'Sin permiso'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No se envió archivo'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({'message': 'Solo se aceptan archivos .xlsx, .xls o .csv'}), 400

    try:
        if file.filename.endswith('.csv'):
            return importar_csv(file)
        return importar_xlsx(file)
    except Exception as e:
        return jsonify({'message': f'Error procesando archivo: {str(e)}'}), 500


def limpiar_precio(valor):
    """Convierte '$1.234,56' o '1234.56' a float."""
    if not valor:
        return None
    s = str(valor).strip().replace('$', '').replace(',', '').replace(' ', '')
    s = s.replace('.', '', s.count('.')-1) if s.count('.') > 1 else s
    try:
        return float(s)
    except:
        return None


def importar_xlsx(file):
    try:
        import openpyxl
    except ImportError:
        return jsonify({'message': 'Instala openpyxl: pip install openpyxl'}), 500

    contenido = file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido))

    # Buscar la hoja con productos
    hoja = None
    for nombre in wb.sheetnames:
        if 'product' in nombre.lower() or 'producto' in nombre.lower() or nombre == wb.sheetnames[0]:
            hoja = wb[nombre]
            break
    if not hoja:
        return jsonify({'message': 'No se encontró hoja de productos'}), 400

    # Leer encabezados
    headers = []
    for cell in hoja[1]:
        val = str(cell.value or '').strip().lower()
        headers.append(val)

    return procesar_filas(hoja, headers, start_row=2, is_xlsx=True)


def importar_csv(file):
    import csv
    contenido = file.read().decode('utf-8-sig')
    reader    = csv.reader(io.StringIO(contenido))
    filas     = list(reader)
    if not filas:
        return jsonify({'message': 'Archivo CSV vacío'}), 400

    headers = [h.strip().lower() for h in filas[0]]
    return procesar_filas(filas, headers, start_row=1, is_xlsx=False)


def get_col(headers, opciones):
    """Busca el índice de una columna por posibles nombres."""
    for op in opciones:
        for i, h in enumerate(headers):
            if op in h:
                return i
    return None


def procesar_filas(data, headers, start_row, is_xlsx):
    # Mapeo de columnas flexible
    col = {
        'nombre':     get_col(headers, ['nombre','name','producto','descripci']),
        'precio':     get_col(headers, ['precio','price','valor','costo']),
        'stock':      get_col(headers, ['stock','cantidad','existencia','inventario']),
        'min_stock':  get_col(headers, ['min','stock min','minimo','minimum']),
        'categoria':  get_col(headers, ['categoria','category','grupo','tipo']),
        'barcode':    get_col(headers, ['codigo','barcode','ean','upc','barra']),
        'proveedor':  get_col(headers, ['proveedor','supplier','marca','brand']),
        'iva':        get_col(headers, ['iva','tax','impuesto']),
        'gramaje_c':  get_col(headers, ['gramaje c','gramaje_c','peso','weight','contenido']),
        'gramaje_u':  get_col(headers, ['gramaje u','gramaje_u','unidad','unit']),
        'vencimiento':get_col(headers, ['vencimiento','expiry','caducidad','fecha v']),
        'descuento':  get_col(headers, ['descuento','discount']),
        'descripcion':get_col(headers, ['descripcion','description','detalle']),
    }

    if col['nombre'] is None:
        return jsonify({'message': 'No se encontró columna "Nombre" en el archivo'}), 400
    if col['precio'] is None:
        return jsonify({'message': 'No se encontró columna "Precio" en el archivo'}), 400

    creados    = 0
    actualizados = 0
    errores    = []

    filas = list(data)[start_row:] if not is_xlsx else list(data.iter_rows(min_row=start_row, values_only=True))

    for num, fila in enumerate(filas, start_row+1):
        try:
            if is_xlsx:
                vals = list(fila)
            else:
                vals = fila

            if not vals or all(v is None or str(v).strip() == '' for v in vals):
                continue

            def get(c):
                if c is None: return None
                if c >= len(vals): return None
                v = vals[c]
                return str(v).strip() if v is not None else None

            nombre = get(col['nombre'])
            precio_raw = get(col['precio'])

            if not nombre or nombre.lower() in ('nombre','name','producto'):
                continue

            precio = limpiar_precio(precio_raw)
            if precio is None or precio <= 0:
                errores.append(f'Fila {num}: precio inválido "{precio_raw}" para "{nombre}"')
                continue

            # Buscar producto existente
            producto = Product.query.filter(
                db.func.lower(Product.name) == nombre.lower()
            ).first()

            if not producto:
                # Buscar por código de barras
                barcode = get(col['barcode'])
                if barcode:
                    producto = Product.query.filter_by(barcode=barcode).first()

            if producto:
                # Actualizar
                producto.price = precio
                if col['stock'] is not None:
                    s = limpiar_precio(get(col['stock']))
                    if s is not None: producto.stock = int(s)
                accion = 'actualizado'
                actualizados += 1
            else:
                # Crear nuevo
                producto = Product(name=nombre, price=precio, is_active=True)
                db.session.add(producto)
                accion = 'creado'
                creados += 1

            # Campos opcionales
            if col['stock'] is not None:
                s = limpiar_precio(get(col['stock']))
                if s is not None: producto.stock = int(s)

            if col['min_stock'] is not None:
                ms = limpiar_precio(get(col['min_stock']))
                if ms is not None: producto.min_stock = int(ms)

            if col['categoria'] is not None:
                cat = get(col['categoria'])
                if cat: producto.category = cat

            barcode = get(col['barcode']) if col['barcode'] is not None else None
            if barcode:
                # Verificar que no esté duplicado
                existe = Product.query.filter_by(barcode=barcode).first()
                if not existe or existe.id == (producto.id if producto.id else 0):
                    producto.barcode = barcode

            if col['iva'] is not None:
                iva = limpiar_precio(get(col['iva']))
                if iva in (0, 5, 19): producto.iva_type = int(iva)

            if col['gramaje_c'] is not None:
                gc = limpiar_precio(get(col['gramaje_c']))
                if gc: producto.gramaje_cantidad = gc

            if col['gramaje_u'] is not None:
                gu = get(col['gramaje_u'])
                if gu: producto.gramaje_unidad = gu.lower()

            if col['descuento'] is not None:
                desc = limpiar_precio(get(col['descuento']))
                if desc is not None: producto.discount = desc

            if col['descripcion'] is not None:
                d = get(col['descripcion'])
                if d: producto.description = d

            # Proveedor por nombre
            if col['proveedor'] is not None:
                prov_nombre = get(col['proveedor'])
                if prov_nombre:
                    prov = Supplier.query.filter(
                        db.or_(
                            db.func.lower(Supplier.name) == prov_nombre.lower(),
                            db.func.lower(Supplier.company_name) == prov_nombre.lower()
                        )
                    ).first()
                    if prov:
                        producto.supplier_id = prov.id

        except Exception as e:
            errores.append(f'Fila {num}: {str(e)}')
            continue

    db.session.commit()

    return jsonify({
        'message':     f'Importación completada: {creados} creados, {actualizados} actualizados',
        'creados':     creados,
        'actualizados':actualizados,
        'errores':     errores,
        'total':       creados + actualizados,
    }), 200


# ── DESCARGAR PLANTILLA ───────────────────────────────────────────────────
@import_export_bp.route('/import/plantilla', methods=['GET'])
@jwt_required()
def descargar_plantilla():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'message': 'Instala openpyxl'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos SmartMerca'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        ('Nombre *',          30), ('Precio *',   12), ('Stock',         10),
        ('Stock Mínimo',      12), ('Categoría',  20), ('Código Barras', 18),
        ('Proveedor',         20), ('IVA (%)',      8), ('Gramaje Cant',  12),
        ('Gramaje Unidad',    12), ('Descuento %', 12), ('Descripción',  30),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Ejemplos
    ejemplos = [
        ['Arroz Diana 500g', 3500, 50, 10, 'Granos',             '7702001234567', 'Diana',   0,  500, 'g',  0,   'Arroz de primera calidad'],
        ['Aceite Palma 1L',  8900, 30, 5,  'Aceites',            '7702009876543', 'Palma',   19, 1,   'l',  0,   ''],
        ['Banano',           2500, 40, 10, 'Frutas y Verduras',  '',              '',        0,  1,   'kg', 0,   ''],
        ['Leche Entera 1L',  3200, 60, 15, 'Lácteos',           '7702005551234', 'Alquería',0,  1,   'l',  0,   ''],
        ['Papa',             2000, 50, 20, 'Frutas y Verduras',  '',              '',        0,  1,   'kg', 0,   ''],
    ]

    alt_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    for row, ej in enumerate(ejemplos, 2):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)
        for col, val in enumerate(ej, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.fill      = fill
            cell.alignment = Alignment(vertical='center')

    # Nota al pie
    ws.cell(row=8, column=1, value='* = Campos obligatorios').font = Font(bold=True, color='FF0000')
    ws.cell(row=9, column=1, value='IVA: 0, 5 o 19').font = Font(italic=True, color='666666')
    ws.cell(row=10, column=1, value='Gramaje Unidad: kg, g, lb, ml, l, unidad').font = Font(italic=True, color='666666')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Plantilla_Importar_Productos.xlsx'
    )